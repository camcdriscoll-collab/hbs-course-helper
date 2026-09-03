#!/usr/bin/env python3
"""
participation_tracker.py — HBS participation tracker spreadsheet.

Creates/refreshes: ~/Desktop/Coursework/Participation Tracker.xlsx
  - Single worksheet ("Participation")
  - CATS | CFO | LME | LTV side by side, each a different color
  - Narrow separator column between each course group
  - Row 1: course name header (dark course color, merged)
  - Row 2: live participation rate "X / Y" (mid course color, merged)
  - Row 3: column labels Day | Case Title | Rating
  - Row 4+: session rows sorted by date

Rating values: ok, good, great, x (didn't speak), or blank (not yet entered)
Denominator = count of non-blank rating cells (entry-based)
Numerator   = count of ok + good + great ratings

On refresh (called from canvas_refresh.py --weekly):
  - Case titles and dates are updated from Canvas
  - User-entered ratings are preserved (keyed by course + session date)

Run standalone:
  ~/repos/hbs-course-helper/.venv/bin/python3 participation_tracker.py
"""

import json
import os
import re
import sys
from datetime import datetime, timezone, timedelta, date as _date
from pathlib import Path
from urllib.error import HTTPError
from urllib.parse import urlencode
from urllib.request import HTTPRedirectHandler, Request, build_opener

# ── Path setup ────────────────────────────────────────────────────────────────

sys.path.insert(0, str(Path(__file__).parent))
import path_config

_paths       = path_config.resolve()
DEST_ROOT    = _paths["coursework_root"]
ENV_FILE     = _paths["env_file"] or Path("/dev/null")
_COURSES     = _paths["courses"]
COURSE_NAMES = path_config.COURSE_NAMES

COURSES      = {a: d["canvas_id"] for a, d in _COURSES.items() if d["folder_path"]}
CANVAS_BASE  = _paths["canvas_base"]
BOSTON       = timezone(timedelta(hours=-4))

COURSE_ORDER    = ["CATS", "CFO", "LME", "LTV"]
COLS_PER_COURSE = 3
STRIDE          = COLS_PER_COURSE + 1   # 4: three data cols + one separator col
DATA_START_ROW  = 4                     # rows 1-3 are header rows
OUTPUT_FILE     = DEST_ROOT / "Participation Tracker.xlsx"

# Per-course color scheme: (header_dark, rate_mid, even_row_light)
COURSE_COLORS = {
    "CATS": ("1E6B4A", "2A9466", "E8F6EF"),   # teal / green
    "CFO":  ("1F4E79", "2E75B6", "EBF3FB"),   # navy / blue
    "LME":  ("7B2133", "B03050", "FAEAED"),   # burgundy / red
    "LTV":  ("4A2178", "6B33A8", "F1ECF9"),   # indigo / purple
}


def course_start_col(i: int) -> int:
    """1-indexed start column for course i: 1, 5, 9, 13."""
    return i * STRIDE + 1


def sep_col_for(i: int) -> int:
    """1-indexed separator column that follows course i (i = 0, 1, 2)."""
    return i * STRIDE + COLS_PER_COURSE + 1   # 4, 8, 12


# ── Canvas API ────────────────────────────────────────────────────────────────


def _load_token() -> str:
    token = os.getenv("CANVAS_API_TOKEN", "")
    if token:
        return token
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text().splitlines():
            line = line.strip()
            if line.startswith("CANVAS_API_TOKEN=") and not line.startswith("#"):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    sys.exit(f"No CANVAS_API_TOKEN found. Check {ENV_FILE}")


class _StripAuth(HTTPRedirectHandler):
    def redirect_request(self, req, fp, code, msg, headers, newurl):
        new_req = super().redirect_request(req, fp, code, msg, headers, newurl)
        if new_req:
            from urllib.parse import urlparse
            if urlparse(newurl).netloc != urlparse(req.full_url).netloc:
                new_req.remove_header("Authorization")
        return new_req


_opener = build_opener(_StripAuth())


def canvas_get(path: str, params: dict | None = None) -> list | dict:
    token = _load_token()
    url = f"{CANVAS_BASE}/{path.lstrip('/')}"
    if params:
        url += "?" + urlencode(params)
    results = []
    while url:
        req = Request(url, headers={"Authorization": f"Bearer {token}"})
        try:
            resp = _opener.open(req, timeout=30)
        except HTTPError as e:
            print(f"    HTTP {e.code}: {url}")
            return []
        data = json.loads(resp.read())
        if isinstance(data, list):
            results.extend(data)
        else:
            return data
        link = resp.headers.get("Link", "")
        url = None
        for part in link.split(","):
            if 'rel="next"' in part:
                m = re.search(r"<(.+?)>", part)
                if m:
                    url = m.group(1)
    return results


# ── Helpers ───────────────────────────────────────────────────────────────────


def boston_date(iso: str) -> datetime:
    return datetime.fromisoformat(iso.replace("Z", "+00:00")).astimezone(BOSTON)


def yymmdd(dt: datetime) -> str:
    return dt.strftime("%y%m%d")


def col_letter(n: int) -> str:
    """1-indexed column number → Excel column letter (A, B, …, Z, AA, …)."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


def extract_case_title(name: str) -> str:
    """
    Strip 'COURSE | Class N: ' boilerplate, return the case/topic title.
    e.g. "CFO | Class 3: The DCF Method" → "The DCF Method"
    """
    m = re.search(r"class\s+\d+\s*[:\-]\s*(.*)", name, re.IGNORECASE)
    if m and m.group(1).strip():
        return m.group(1).strip()
    if "|" in name:
        return name.split("|")[-1].strip()
    return name.strip()


# ── Session data ──────────────────────────────────────────────────────────────


def get_all_sessions() -> dict[str, list[dict]]:
    """
    Fetch all Canvas assignments with due dates for each active course.
    Returns {abbrev: [assignment_dict, ...]}, sorted chronologically.
    """
    result: dict[str, list[dict]] = {}
    for abbrev in COURSE_ORDER:
        course_id = COURSES.get(abbrev)
        if not course_id:
            result[abbrev] = []
            continue
        assignments = canvas_get(f"courses/{course_id}/assignments", {"per_page": 100})
        sessions = [a for a in assignments if a.get("due_at")]
        sessions.sort(key=lambda a: a["due_at"])
        result[abbrev] = sessions
    return result


# ── Preserve existing ratings ─────────────────────────────────────────────────


def read_existing_ratings(path: Path) -> dict[str, dict[str, str]]:
    """
    Read user-entered ratings from an existing spreadsheet.
    Returns {abbrev: {yymmdd_key: rating_string}}.
    Scans from row 3 onward and identifies cells by actual date values,
    so it survives both the old and new header row layout.
    """
    ratings: dict[str, dict[str, str]] = {a: {} for a in COURSE_ORDER}
    try:
        import openpyxl
    except ImportError:
        return ratings

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        for i, abbrev in enumerate(COURSE_ORDER):
            # Use current column layout (with separator stride)
            day_col    = course_start_col(i)
            rating_col = course_start_col(i) + 2

            for row in ws.iter_rows(min_row=3):
                if len(row) < rating_col:
                    continue
                day_val    = row[day_col - 1].value
                rating_val = row[rating_col - 1].value

                if day_val is None:
                    continue
                if isinstance(day_val, datetime):
                    day_val = day_val.date()
                if not isinstance(day_val, _date):
                    continue

                key    = day_val.strftime("%y%m%d")
                rating = str(rating_val or "").strip().lower()
                if rating:
                    ratings[abbrev][key] = rating
    except Exception as e:
        print(f"  Warning: could not read existing ratings ({e}) — starting fresh")

    return ratings


# ── Spreadsheet builder ───────────────────────────────────────────────────────


def build_tracker():
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        sys.exit(
            "openpyxl not installed.\n"
            "Run: ~/repos/hbs-course-helper/.venv/bin/python3 -m pip install openpyxl"
        )

    # ── Preserve existing user ratings ────────────────────────────────────────
    existing: dict[str, dict[str, str]] = {a: {} for a in COURSE_ORDER}
    if OUTPUT_FILE.exists():
        existing = read_existing_ratings(OUTPUT_FILE)
        total = sum(len(v) for v in existing.values())
        if total:
            print(f"  Preserved {total} existing rating(s).")

    # ── Fetch Canvas sessions ─────────────────────────────────────────────────
    print("  Fetching sessions from Canvas...")
    sessions = get_all_sessions()
    for abbrev in COURSE_ORDER:
        print(f"    {abbrev}: {len(sessions.get(abbrev, []))} session(s)")

    max_sessions = max((len(v) for v in sessions.values()), default=0)
    # Formula ranges extend a bit past the last data row for future sessions
    formula_end = DATA_START_ROW + max(max_sessions, 50) + 10

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participation"

    WHITE  = "FFFFFF"
    center   = Alignment(horizontal="center", vertical="center")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    # ── Column widths ─────────────────────────────────────────────────────────
    # Day=11, Case Title=42, Rating=10 per course; separator=2
    col_widths = [11, 42, 10]
    for i in range(len(COURSE_ORDER)):
        sc = course_start_col(i)
        for j, w in enumerate(col_widths):
            ws.column_dimensions[col_letter(sc + j)].width = w
        if i < len(COURSE_ORDER) - 1:
            ws.column_dimensions[col_letter(sep_col_for(i))].width = 2

    # ── Row heights ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24   # course name
    ws.row_dimensions[2].height = 22   # participation rate
    ws.row_dimensions[3].height = 18   # column labels

    # ── Per-course headers (rows 1–3) ─────────────────────────────────────────
    for i, abbrev in enumerate(COURSE_ORDER):
        sc        = course_start_col(i)
        ec        = sc + COLS_PER_COURSE - 1
        rating_c  = col_letter(sc + 2)   # A+2, E+2 etc → C, G, K, O
        full_name = COURSE_NAMES.get(abbrev, abbrev)

        dark, mid, _ = COURSE_COLORS[abbrev]
        dark_fill = PatternFill(fill_type="solid", fgColor=dark)
        mid_fill  = PatternFill(fill_type="solid", fgColor=mid)
        bold_white = Font(bold=True, color=WHITE, size=12)

        # ── Row 1: Course name ────────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        c1 = ws.cell(row=1, column=sc, value=full_name)
        c1.font      = bold_white
        c1.fill      = dark_fill
        c1.alignment = center

        # ── Row 2: Live participation rate "spoke / entered" ──────────────────
        num_f = (
            f"SUMPRODUCT(({rating_c}{DATA_START_ROW}:{rating_c}{formula_end}=\"ok\")"
            f"+({rating_c}{DATA_START_ROW}:{rating_c}{formula_end}=\"good\")"
            f"+({rating_c}{DATA_START_ROW}:{rating_c}{formula_end}=\"great\"))"
        )
        den_f = (
            f"SUMPRODUCT(({rating_c}{DATA_START_ROW}:{rating_c}{formula_end}<>\"\"))"
        )
        rate_formula = f'=TEXT({num_f},"0")&" / "&TEXT({den_f},"0")'

        ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        c2 = ws.cell(row=2, column=sc, value=rate_formula)
        c2.font      = bold_white
        c2.fill      = mid_fill
        c2.alignment = center

        # ── Row 3: Column labels ──────────────────────────────────────────────
        for j, label in enumerate(["Day", "Case Title", "Rating"]):
            c3 = ws.cell(row=3, column=sc + j, value=label)
            c3.font      = bold_white
            c3.fill      = dark_fill
            c3.alignment = center

    # ── Data rows (row 4+) ────────────────────────────────────────────────────
    for i, abbrev in enumerate(COURSE_ORDER):
        sc = course_start_col(i)
        _, _, light = COURSE_COLORS[abbrev]
        even_fill = PatternFill(fill_type="solid", fgColor=light)

        course_sessions = sessions.get(abbrev, [])
        existing_abbrev = existing.get(abbrev, {})

        for row_offset, a in enumerate(course_sessions):
            row      = DATA_START_ROW + row_offset
            dt       = boston_date(a["due_at"])
            date_val = dt.date()
            date_key = yymmdd(dt)
            title    = extract_case_title(a.get("name", ""))
            rating   = existing_abbrev.get(date_key, "")

            bg = even_fill if (row_offset % 2 == 1) else None

            # Day
            day_cell = ws.cell(row=row, column=sc, value=date_val)
            day_cell.number_format = "MMM D"
            day_cell.alignment     = center
            day_cell.font          = Font(size=12)
            if bg:
                day_cell.fill = bg

            # Case title
            title_cell = ws.cell(row=row, column=sc + 1, value=title)
            title_cell.alignment = wrap_top
            title_cell.font      = Font(size=12)
            if bg:
                title_cell.fill = bg

            # Rating (user-filled)
            rating_cell = ws.cell(row=row, column=sc + 2, value=rating)
            rating_cell.alignment = center
            rating_cell.font      = Font(size=12)
            if bg:
                rating_cell.fill = bg

        # ── Dropdown validation for Rating column ─────────────────────────────
        rating_col_letter = col_letter(sc + 2)
        dv_end = DATA_START_ROW + max(len(course_sessions), 40)
        dv = DataValidation(
            type="list",
            formula1='"ok,good,great,x"',
            allow_blank=True,
            showDropDown=False,
            error="Enter ok, good, great, or x",
            errorTitle="Invalid entry",
            prompt="ok = brief comment\ngood = solid contribution\ngreat = strong insight\nx = didn't speak",
            promptTitle="Participation",
        )
        ws.add_data_validation(dv)
        dv.sqref = f"{rating_col_letter}{DATA_START_ROW}:{rating_col_letter}{dv_end}"

    # ── Conditional formatting: color-code ratings ────────────────────────────
    # great = green  good = light green  ok = yellow  x = gray
    rating_colors = [
        ("great", "C6EFCE", "375623", True),
        ("good",  "E2EFDA", "375623", False),
        ("ok",    "FFEB9C", "9C5700", False),
        ("x",     "F2F2F2", "7F7F7F", False),
    ]
    for i in range(len(COURSE_ORDER)):
        sc       = course_start_col(i)
        rating_c = col_letter(sc + 2)
        rng      = f"{rating_c}{DATA_START_ROW}:{rating_c}{formula_end}"

        for value, bg, fg, bold in rating_colors:
            fill = PatternFill(fill_type="solid", fgColor=bg)
            font = Font(color=fg, bold=bold, size=12)
            rule = CellIsRule(operator="equal", formula=[f'"{value}"'],
                              fill=fill, font=font)
            ws.conditional_formatting.add(rng, rule)

    # ── Freeze top 3 header rows ──────────────────────────────────────────────
    ws.freeze_panes = f"A{DATA_START_ROW}"

    # ── Save ──────────────────────────────────────────────────────────────────
    try:
        wb.save(OUTPUT_FILE)
    except PermissionError:
        sys.exit(
            f"\nCould not save — is {OUTPUT_FILE.name} open in Excel? "
            "Close it and try again."
        )
    print(f"  Saved: {OUTPUT_FILE}")


# ── Public entry point (called from canvas_refresh.py) ───────────────────────


def refresh():
    """Refresh the participation tracker. Called from canvas_refresh.py --weekly."""
    build_tracker()


# ── Standalone ───────────────────────────────────────────────────────────────


def main():
    print("Building HBS participation tracker...")
    build_tracker()
    print("Done.")


if __name__ == "__main__":
    main()
